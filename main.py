import webbrowser
import requests
import urllib.request
import json
import re
import time
import ipaddress
from bs4 import BeautifulSoup as BS
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import  Workbook, load_workbook
from requests.packages.urllib3.exceptions import InsecureRequestWarning
from tkinter.filedialog import askopenfilename
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# driver=webdriver.Chrome()
# driver.get('https://www.similarweb.com/website/brickapic.com/#overview')
# time.sleep(7)
# driver.execute_script("window.stop();")
# time.sleep(3)
# driver.delete_all_cookies()
# driver.get('https://www.similarweb.com/website/everestmemory.com/#overview')
# time.sleep(5)
# print(driver.page_source)

df= pd.read_excel('C:\\Users\\semch\\PycharmProjects\\pqm\\data_pqm.xlsx',sheet_name='Table3 (events)')
print (df.head())
type_cods = {
"003001000000" : "ЗЕМЛИ СЕЛЬСКОХОЗЯЙСТВЕННОГО НАЗНАЧЕНИЯ",
"003001000010" : "Сельскохозяйственные угодья",
"003001000020" : "Земельные участки, занятые внутрихозяйственными дорогами, коммуникациями, лесными насаждениями, предназначенными для обеспечения защиты земель от негативного воздействия, водными объектами, а также занятые зданиями, сооружениями, используемыми для производства, хранения и первичной переработки сельскохозяйственной продукции",
"003001000030" : "Прочие земельные участки из состава земель сельскохозяйственного назначения",
"003002000000" : "Земли населенных пунктов",
"003002000010" : "Земельные участки, отнесенные к зонам сельскохозяйственного использования",
"003002000020" : "Земельные участки, занятые жилищным фондом и объектами инженерной инфраструктуры жилищно-коммунального комплекса",
"003002000030" : "Земельные участки, приобретенные (предоставленные) для индивидуального жилищного строительства",
"003002000040" : "Земельные участки, приобретенные (предоставленные) на условиях осуществления на них жилищного строительства (за исключением индивидуального жилищного строительства)",
"003002000060" : "Земельные участки, приобретенные (предоставленные) для ведения личного подсобного хозяйства, садоводства и огородничества или животноводства, а также дачного хозяйства",
"003002000090" : "Земельные участки, отнесенные к производственным территориальным зонам и зонам инженерных и транспортных инфраструктур",
"003002000110" : "Земельные участки для обеспечения обороны",
"003002000120" : "Земельные участки для обеспечения безопасности",
"003002000130" : "Земельные участки для обеспечения таможенных нужд",
"003002000100" : "Прочие земельные участки",
"003003000000" : "Земли промышленности, энергетики, транспорта, связи, радиовещания, телевидения, информатики, земли для обеспечения космической деятельности, земли обороны, безопасности и земли иного специального назначения",
"003003000010" : "Земельные участки из состава земель промышленности",
"003003000020" : "Земельные участки из состава земель энергетики",
"003003000030" : "Земельные участки из состава земель транспорта",
"003003000040" : "Земельные участки из состава земель связи, радиовещания, телевидения, информатики",
"003003000060" : "Земельные участки из состава земель обороны",
"003003000070" : "Земельные участки из состава земель безопасности",
"003008000010" : "Земельные участки из состава земель для обеспечения таможенных нужд",
"003003000080" : "Земельные участки из состава земель иного специального назначения",
"003004000000" : "Земли особо охраняемых территорий и объектов",
"003005000000" : "Земли лесного фонда",
"003006000000" : "Земли водного фонда",
"003007000000" : "Земли запаса",
"003008000000" : "Земельные участки, для которых категория земель не установлена"

}

try:
    print('ВЫ ЗАПУСТИЛИ ПРОГРАММУ ПО СБОРУ ДАННЫХ ИЗ РОСРЕЕСТРА (pkk.ru)\n'
          'Для корректной работы программы необходимо выбрать Excel-файл, в котором есть лист с названием "pkk.rosreestr"\n'
          'на котором ячейка "А1" имеет значение "Кадастровый номер", а ниже, в этом же столбце перечислены кадастровые номера.\n'
          '\n\n\nНажмите Enter, чтобы выбрать файл "Excel"')

    input()

    filename = askopenfilename()
    requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

    while filename == '':
        print('Вы не выбрали файл')
        filename = askopenfilename()


    book = load_workbook(filename=filename)
    sheet = book['pkk.rosreestr']
    if(sheet['A1'].value != 'Кадастровый номер'):
        input('ОШИБКА. Вы выбрали файл в котором ячейка А1 не содержит "Кадастровый номер" \n\nЧтобы выйти из программы нажмите Enter...')
        exit()
    sheet['B1'] = 'Категория земель'
    sheet['C1'] = 'Разрешенное использование'
    sheet['D1'] = 'Площадь'
    sheet['E1'] = 'Кадастровая стоимость'
    sheet['F1'] = 'Адрес'

    # data = requests.get('https://pkk.rosreestr.ru/api/features/1/50:8:70226:470', verify=False)
    # output = json.loads(data.text)
    # cost = output['feature']['attrs']['cad_cost']
    # print (cost)
    count = sheet.max_row+1
    row = 2
    countempty = 0

    while countempty <= 10:
        retries = 1
        success = False
        while not success or countempty <= 10 :
            try:
                if(sheet[row][0].value == None):
                    countempty += 1
                    row += 1
                    continue
                cad_to_link = sheet[row][0].value
                cad_ip = re.sub('[:]0+', ':', cad_to_link)
                cad = re.sub(r'::', ':0:', cad_ip)
                cad = cad.strip()
                data = requests.get('https://pkk.rosreestr.ru/api/features/1/' + cad, verify=False, timeout=(30,10))
                output = json.loads(data.text)
                if (output['feature'] == None):
                    sheet[row][1].value = 'НЕДЕЙСТВИТЕЛЬНЫЙ КАД.НОМЕР'
                    print('Кадастровый номер в строке %s не является действительным' %row)
                    row += 1
                    continue
                square = output['feature']['attrs']['area_value']
                sheet[row][3].value = square

                cost = output['feature']['attrs']['cad_cost']
                sheet[row][4].value = cost
                try:
                    land_category_code = output['feature']['attrs']['category_type']
                    land_category = type_cods[land_category_code]
                    sheet[row][1].value = land_category
                except:
                    print("Ошибка кода категории земель")
                    sheet[row][1].value = 'НЕ ОПРЕДЕЛЕН'
                    pass
                allowed_using = output['feature']['attrs']['util_by_doc']
                sheet[row][2].value = allowed_using
                address = output['feature']['attrs']['address']

                if (address == None):
                    sheet[row][5].value = 'НЕ УСТАНОВЛЕН'
                else:
                    sheet[row][5].value = address
                print(row, square, cost, address, sep='  ----  ')
                row += 1
                countempty = 0
                success = True
            except Exception as e:
                wait = retries * 2
                print('Ожидайте, программа выполняется')
                #time.sleep(wait)
                retries += 1



    book.save(filename)
    book.close
    input("Программа завершена успешно, нажите Enter, чтобы выйти ...")
except Exception as e:
    print(e)
    input('Программа завершила свою работу с ошибкой, нажмите Enter чтобы выйти ...')












